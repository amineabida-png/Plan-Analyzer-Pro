"""
Exportateur Excel du métré (Plan Analyzer Pro).
Produit un classeur professionnel à partir d'un RapportAnalyse :
  - feuille "Métré" : postes, quantités, unités, prix unitaire (à saisir), total (formule)
  - feuille "Synthèse" : totaux par lot avec formules
Les colonnes Prix Unitaire sont en bleu (saisie utilisateur), les totaux en formule.
"""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import RapportAnalyse

# Police professionnelle imposée par les bonnes pratiques
POLICE = "Arial"
BLEU_SAISIE = Font(name=POLICE, color="0000FF")       # entrées modifiables
NOIR = Font(name=POLICE, color="000000")
ENTETE = Font(name=POLICE, bold=True, color="FFFFFF")
FILL_ENTETE = PatternFill("solid", start_color="1F4E78")
FILL_TITRE = PatternFill("solid", start_color="D9E1F2")
BORDURE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
CENTRE = Alignment(horizontal="center", vertical="center")


def exporter_excel(rapport: RapportAnalyse, chemin: str) -> str:
    wb = Workbook()

    # ---------------- Feuille Métré ----------------
    ws = wb.active
    ws.title = "Métré"

    ws["A1"] = "Plan Analyzer Pro — MÉTRÉ AUTOMATIQUE"
    ws["A1"].font = Font(name=POLICE, bold=True, size=14)
    ws["A2"] = f"Fichier : {rapport.fichier}"
    ws["A3"] = f"Unité de dessin détectée : {rapport.unite_dessin.value}"
    ws["A2"].font = ws["A3"].font = Font(name=POLICE, italic=True, size=9)

    entetes = ["N°", "Poste", "Détail", "Quantité", "Unité",
               "Prix Unitaire (€)", "Montant (€)"]
    ligne_entete = 5
    for col, titre in enumerate(entetes, start=1):
        c = ws.cell(row=ligne_entete, column=col, value=titre)
        c.font = ENTETE
        c.fill = FILL_ENTETE
        c.alignment = CENTRE
        c.border = BORDURE

    r = ligne_entete + 1
    premiere_ligne_data = r
    for i, l in enumerate(rapport.metre, start=1):
        ws.cell(row=r, column=1, value=i).font = NOIR
        ws.cell(row=r, column=2, value=l.poste).font = NOIR
        ws.cell(row=r, column=3, value=l.detail or "").font = NOIR
        ws.cell(row=r, column=4, value=l.quantite).font = NOIR
        ws.cell(row=r, column=5, value=l.unite).font = NOIR
        # Prix unitaire : saisie utilisateur (bleu), 0 par défaut
        pu = ws.cell(row=r, column=6, value=0)
        pu.font = BLEU_SAISIE
        # Montant = Quantité * PU (FORMULE, jamais codé en dur)
        mt = ws.cell(row=r, column=7, value=f"=D{r}*F{r}")
        mt.font = NOIR
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = BORDURE
        r += 1
    derniere_ligne_data = r - 1

    # Ligne total
    ws.cell(row=r, column=2, value="TOTAL GÉNÉRAL HT").font = Font(name=POLICE, bold=True)
    total = ws.cell(row=r, column=7,
                    value=f"=SUM(G{premiere_ligne_data}:G{derniere_ligne_data})")
    total.font = Font(name=POLICE, bold=True)
    for col in range(2, 8):
        ws.cell(row=r, column=col).fill = FILL_TITRE

    # Formats de nombres
    for row in range(premiere_ligne_data, derniere_ligne_data + 1):
        ws.cell(row=row, column=4).number_format = "#,##0.00"
        ws.cell(row=row, column=6).number_format = "#,##0.00 €"
        ws.cell(row=row, column=7).number_format = "#,##0.00 €"
    ws.cell(row=r, column=7).number_format = "#,##0.00 €"

    # Largeurs de colonnes
    largeurs = {"A": 5, "B": 40, "C": 28, "D": 12, "E": 8, "F": 18, "G": 16}
    for col, w in largeurs.items():
        ws.column_dimensions[col].width = w

    # ---------------- Feuille Pièces ----------------
    if rapport.pieces:
        wsp = wb.create_sheet("Pièces")
        wsp["A1"] = "DÉTAIL DES PIÈCES"
        wsp["A1"].font = Font(name=POLICE, bold=True, size=13)
        entetes_p = ["Pièce", "Surface (m²)", "Périmètre (m)",
                     "Carrelage (m²)", "Plafond (m²)", "Peinture murs (m²)"]
        lp = 3
        for col, titre in enumerate(entetes_p, start=1):
            c = wsp.cell(row=lp, column=col, value=titre)
            c.font = ENTETE
            c.fill = FILL_ENTETE
            c.alignment = CENTRE
            c.border = BORDURE
        r = lp + 1
        debut = r
        for p in rapport.pieces:
            wsp.cell(row=r, column=1, value=p.nom).font = NOIR
            wsp.cell(row=r, column=2, value=p.surface_m2).font = NOIR
            wsp.cell(row=r, column=3, value=p.perimetre_m).font = NOIR
            wsp.cell(row=r, column=4, value=p.surface_carrelage_m2).font = NOIR
            wsp.cell(row=r, column=5, value=p.surface_plafond_m2).font = NOIR
            wsp.cell(row=r, column=6, value=p.surface_peinture_murs_m2).font = NOIR
            for col in range(1, 7):
                cell = wsp.cell(row=r, column=col)
                cell.border = BORDURE
                if col >= 2:
                    cell.number_format = "#,##0.00"
            r += 1
        # Totaux en formules
        wsp.cell(row=r, column=1, value="TOTAL").font = Font(name=POLICE, bold=True)
        for col, lettre in [(2, "B"), (4, "D"), (5, "E"), (6, "F")]:
            t = wsp.cell(row=r, column=col,
                         value=f"=SUM({lettre}{debut}:{lettre}{r-1})")
            t.font = Font(name=POLICE, bold=True)
            t.number_format = "#,##0.00"
            t.fill = FILL_TITRE
        for col, w in zip("ABCDEF", [20, 13, 14, 14, 13, 17]):
            wsp.column_dimensions[col].width = w

    # ---------------- Feuille Alertes ----------------
    if rapport.alertes:
        wsa = wb.create_sheet("Alertes")
        wsa["A1"] = "ALERTES À VALIDER PAR LE MÉTREUR"
        wsa["A1"].font = Font(name=POLICE, bold=True, size=12, color="C00000")
        for idx, a in enumerate(rapport.alertes, start=3):
            wsa.cell(row=idx, column=1, value=f"• {a}").font = NOIR
        wsa.column_dimensions["A"].width = 90

    wb.save(chemin)
    return chemin
